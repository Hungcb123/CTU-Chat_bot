#!/usr/bin/env python3
"""
Ablation Test Runner — Table 6: Graph & Agent Ablations
========================================================
Đo lường hiệu suất của các variant kiến trúc:
  1. fixed_route  + hybrid: Rule-based _classify_one() → RAG → LLM
  2. supervisor   + hybrid: LLM RouteDecision → RAG → LLM  
  3. full_system:           Full multi-agent graph (ainvoke)

Metrics:
  - Route accuracy:  actual_route matches expected_path
  - Tool validity:   correct tool was called with valid args
  - Answer contains: expected_contains ⊂ response
  - Answer excludes: expected_not_contains ∩ response = ∅

Dataset: data/test_dataset_rag.json (30 cases)

Usage:
    # Dry-run (3 cases only, no LLM calls for route-only tests)
    python scripts/run_ablation_test.py --variant fixed_route --dry-run

    # Full run for a specific variant
    python scripts/run_ablation_test.py --variant supervisor

    # Run all variants  
    python scripts/run_ablation_test.py --all

    # Run all and export Excel
    python scripts/run_ablation_test.py --all --export-excel
"""

from __future__ import annotations

import argparse
import asyncio
import json
import logging
import os
import sys
import time
from collections import defaultdict
from dataclasses import dataclass, field, asdict
from pathlib import Path
from typing import Any

# Project root
ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
)
logger = logging.getLogger("ablation")

DATASET_PATH = ROOT / "data" / "test_dataset_rag.json"
RESULTS_DIR = ROOT / "data" / "Intelligent-Multi-Agent-System-dev_quynh" / "test_experiment" / "file_test" / "result_test_3_team"


# ─────────────────────────────────────────────────────────────────────
# DATA MODEL
# ─────────────────────────────────────────────────────────────────────

@dataclass
class TestCase:
    id: int
    category: str
    query: str
    expected_contains: list[str]
    expected_not_contains: list[str]
    source_files: list[str]
    expected_path: str
    expected_tool: str | None


@dataclass
class CaseResult:
    case_id: int
    variant: str
    category: str
    query: str

    # Route accuracy
    expected_route: str = ""
    actual_route: str = ""
    route_correct: bool | None = None  # None = n/a

    # Tool validity
    expected_tool: str | None = None
    actual_tool: str | None = None
    tool_correct: bool | None = None

    # Answer metrics
    contains_hits: int = 0
    contains_total: int = 0
    contains_pass: bool = False
    excludes_hits: int = 0
    excludes_total: int = 0
    excludes_pass: bool = True

    # Raw data
    response: str = ""
    error: str = ""
    latency_ms: int = 0


# ─────────────────────────────────────────────────────────────────────
# DATASET LOADER
# ─────────────────────────────────────────────────────────────────────

def load_dataset(path: Path = DATASET_PATH, limit: int | None = None) -> list[TestCase]:
    raw = json.loads(path.read_text(encoding="utf-8"))
    cases = []
    for item in raw:
        cases.append(TestCase(
            id=item["id"],
            category=item.get("category", ""),
            query=item["query"],
            expected_contains=item.get("expected_contains", []),
            expected_not_contains=item.get("expected_not_contains", []),
            source_files=item.get("source_files", []),
            expected_path=item.get("expected_path", ""),
            expected_tool=item.get("expected_tool"),
        ))
    if limit:
        cases = cases[:limit]
    logger.info("Loaded %d test cases from %s", len(cases), path.name)
    return cases


# ─────────────────────────────────────────────────────────────────────
# METRICS
# ─────────────────────────────────────────────────────────────────────

def check_answer_contains(response: str, expected: list[str]) -> tuple[int, int]:
    """Returns (hits, total)."""
    hits = sum(1 for e in expected if e.lower() in response.lower())
    return hits, len(expected)


def check_answer_excludes(response: str, excluded: list[str]) -> tuple[int, int]:
    """Returns (violations, total). 0 violations = pass."""
    violations = sum(1 for e in excluded if e.lower() in response.lower())
    return violations, len(excluded)


# ─────────────────────────────────────────────────────────────────────
# INTENT → ROUTE MAPPING
# ─────────────────────────────────────────────────────────────────────

# Map expected_path from dataset to the agent route name
PATH_TO_ROUTE = {
    "structured_lookup": "financial",
    "structured_lookup_clarification": "financial",
    "rag_exemption_basis": "financial",
    "rag_exemption_policy": "financial",
    "calculation_tool": "financial",
    "rag_scholarship": "scholarship",
    "scholarship_tool": "scholarship",
    "rag_student_loan": "general",
    "rag_general": "general",
    "academic_program": "academic",
}


# ─────────────────────────────────────────────────────────────────────
# VARIANT 1: fixed_route (NO LLM, rule-based only)
# ─────────────────────────────────────────────────────────────────────

def run_fixed_route(cases: list[TestCase]) -> list[CaseResult]:
    """Test deterministic routing accuracy without any LLM calls."""
    from app.services.query_intent import (
        QueryIntent,
        classify_query_intent,
        build_retrieval_lanes,
    )

    # Map QueryIntent → agent route
    INTENT_TO_ROUTE = {
        QueryIntent.ACTUAL_TUITION: "financial",
        QueryIntent.AMBIGUOUS_TUITION: "financial",
        QueryIntent.EXEMPTION_BASIS: "financial",
        QueryIntent.EXEMPTION_POLICY: "financial",
        QueryIntent.CALCULATION: "financial",
        QueryIntent.BOTH: "financial",
        QueryIntent.SCHOLARSHIP: "scholarship",
        QueryIntent.STUDENT_LOAN: "general",
        QueryIntent.SOCIAL_SUPPORT: "general",
        QueryIntent.ACADEMIC_PROGRAM: "academic",
        QueryIntent.ACADEMIC_RULES: "academic",
        QueryIntent.QUY_CHE_GENERAL: "general",
        QueryIntent.OTHER: "general",
    }

    results = []
    for case in cases:
        t0 = time.time()
        decision = classify_query_intent(case.query)
        latency = int((time.time() - t0) * 1000)

        actual_route = INTENT_TO_ROUTE.get(decision.intent, "general")
        expected_route = PATH_TO_ROUTE.get(case.expected_path, "general")

        # Get retrieval lane names for graph/path metric
        lanes = build_retrieval_lanes(decision)
        lane_names = [l.name for l in lanes]

        result = CaseResult(
            case_id=case.id,
            variant="fixed_route",
            category=case.category,
            query=case.query,
            expected_route=expected_route,
            actual_route=actual_route,
            route_correct=(actual_route == expected_route),
            expected_tool=case.expected_tool,
            actual_tool=None,  # No tool calling in this variant
            tool_correct=None,
            response=f"[route={actual_route}, intent={decision.intent.value}, lanes={lane_names}]",
            latency_ms=latency,
        )
        results.append(result)
        logger.debug("Case %d: route=%s (expected=%s) ✓=%s",
                      case.id, actual_route, expected_route, result.route_correct)

    return results


# ─────────────────────────────────────────────────────────────────────
# VARIANT 2: supervisor (LLM routing only, no full answer generation)
# ─────────────────────────────────────────────────────────────────────

async def run_supervisor(cases: list[TestCase]) -> list[CaseResult]:
    """Test LLM Supervisor routing accuracy."""
    from dotenv import load_dotenv
    load_dotenv(ROOT / ".env")

    from langchain_google_genai import ChatGoogleGenerativeAI
    from langchain_core.messages import HumanMessage, SystemMessage
    from pydantic import BaseModel, Field
    from typing import Literal

    from app.agents.prompts import SUPERVISOR_PROMPT

    class RouteDecision(BaseModel):
        next_agent: Literal["academic", "financial", "scholarship", "general"] = Field(
            description="Tên agent chuyên môn sẽ xử lý câu hỏi này"
        )

    llm = ChatGoogleGenerativeAI(model="gemini-3.5-flash-lite")
    supervisor_llm = llm.with_structured_output(RouteDecision)

    results = []
    for case in cases:
        t0 = time.time()
        messages = [
            SystemMessage(content=SUPERVISOR_PROMPT),
            HumanMessage(content=case.query),
        ]
        try:
            route: RouteDecision = await supervisor_llm.ainvoke(messages)
            actual_route = route.next_agent
            error = ""
        except Exception as e:
            actual_route = "general"
            error = str(e)

        latency = int((time.time() - t0) * 1000)
        expected_route = PATH_TO_ROUTE.get(case.expected_path, "general")

        result = CaseResult(
            case_id=case.id,
            variant="supervisor",
            category=case.category,
            query=case.query,
            expected_route=expected_route,
            actual_route=actual_route,
            route_correct=(actual_route == expected_route),
            expected_tool=case.expected_tool,
            tool_correct=None,
            response=f"[route={actual_route}]",
            error=error,
            latency_ms=latency,
        )
        results.append(result)
        status = "✅" if result.route_correct else "❌"
        logger.info("Case %d: %s route=%s (expected=%s) [%dms]",
                     case.id, status, actual_route, expected_route, latency)

        # Rate limit: free tier = 15 req/min → wait 4s between calls
        if case is not cases[-1]:
            await asyncio.sleep(4)

    return results


# ─────────────────────────────────────────────────────────────────────
# VARIANT 3: full_system (complete multi-agent graph)
# ─────────────────────────────────────────────────────────────────────

async def run_full_system(cases: list[TestCase]) -> list[CaseResult]:
    """Test the full multi-agent system end-to-end."""
    from dotenv import load_dotenv
    load_dotenv(ROOT / ".env")

    from langchain_google_genai import ChatGoogleGenerativeAI
    from app.services.rag_engine import AdvancedChunkingEngine
    from app.services.tuition_catalog import TuitionRateCatalog
    from app.services.graph_service import AcademicGraphService
    from app.tools.scholarship import tinh_tien_hoc_bong
    from app.tools.tuition import tinh_toan_hoc_phi
    from app.tools.academic_program import (
        tra_cuu_nganh, so_sanh_nganh, tim_nganh,
        xem_chuoi_tien_quyet, mon_chung_giua_nganh, tim_nganh_co_mon,
        set_graph_service,
    )
    from app.tools.tuition_graph import (
        tra_cuu_hoc_phi_graph, tra_cuu_quy_dinh_hoc_phi,
        set_tuition_graph_service, set_tuition_catalog,
    )
    from app.agents.graph import build_agent_graph

    # Initialize services
    logger.info("🔧 Initializing full system services...")

    llm = ChatGoogleGenerativeAI(model="gemini-3.5-flash-lite")
    rewrite_llm = ChatGoogleGenerativeAI(model="gemini-3.1-flash-lite", temperature=0.0)
    engine = AdvancedChunkingEngine(persist_dir=str(ROOT / "qdrant_storage"))
    tuition_catalog = TuitionRateCatalog.load()

    # Neo4j
    neo4j_uri = os.environ.get("NEO4J_URI", "bolt://localhost:7687")
    neo4j_user = os.environ.get("NEO4J_USER", "neo4j")
    neo4j_password = os.environ.get("NEO4J_PASSWORD", "password")
    graph_service = AcademicGraphService(
        uri=neo4j_uri, user=neo4j_user, password=neo4j_password,
    )
    graph_service.ensure_data_loaded()
    set_graph_service(graph_service)
    set_tuition_graph_service(graph_service)
    set_tuition_catalog(tuition_catalog)

    academic_tools = [
        tra_cuu_nganh, so_sanh_nganh, tim_nganh,
        xem_chuoi_tien_quyet, mon_chung_giua_nganh, tim_nganh_co_mon,
    ]
    financial_tools = [tra_cuu_hoc_phi_graph, tra_cuu_quy_dinh_hoc_phi, tinh_toan_hoc_phi]
    scholarship_tools = [tinh_tien_hoc_bong]

    agent_graph = build_agent_graph(
        llm=llm,
        rewrite_llm=rewrite_llm,
        engine=engine,
        tuition_catalog=tuition_catalog,
        graph_service=graph_service,
        academic_tools=academic_tools,
        financial_tools=financial_tools,
        scholarship_tools=scholarship_tools,
    )

    logger.info("✅ Full system initialized. Running %d cases...", len(cases))

    results = []
    for case in cases:
        t0 = time.time()
        try:
            state = await agent_graph.ainvoke({
                "query": case.query,
                "chat_history": [],
            })
            response = state.get("response", "")
            actual_route = state.get("next_agent", "unknown")
            error = ""
        except Exception as e:
            response = ""
            actual_route = "error"
            error = f"{type(e).__name__}: {e}"
            logger.error("Case %d failed: %s", case.id, error)

        latency = int((time.time() - t0) * 1000)
        expected_route = PATH_TO_ROUTE.get(case.expected_path, "general")

        # Answer metrics
        c_hits, c_total = check_answer_contains(response, case.expected_contains)
        e_violations, e_total = check_answer_excludes(response, case.expected_not_contains)

        result = CaseResult(
            case_id=case.id,
            variant="full_system",
            category=case.category,
            query=case.query,
            expected_route=expected_route,
            actual_route=actual_route,
            route_correct=(actual_route == expected_route),
            expected_tool=case.expected_tool,
            tool_correct=None,  # TODO: extract from agent trace
            contains_hits=c_hits,
            contains_total=c_total,
            contains_pass=(c_hits == c_total),
            excludes_hits=e_violations,
            excludes_total=e_total,
            excludes_pass=(e_violations == 0),
            response=response[:500],
            error=error,
            latency_ms=latency,
        )
        results.append(result)

        status = "✅" if result.contains_pass and result.excludes_pass else "❌"
        logger.info(
            "Case %d: %s route=%s contains=%d/%d excludes=%d/%d [%dms]",
            case.id, status, actual_route,
            c_hits, c_total, e_violations, e_total, latency,
        )

    # Cleanup
    graph_service.close()
    return results


# ─────────────────────────────────────────────────────────────────────
# SUMMARY / REPORTING
# ─────────────────────────────────────────────────────────────────────

def summarize(results: list[CaseResult]) -> dict:
    """Build summary dict grouped by variant."""
    by_variant: dict[str, list[CaseResult]] = defaultdict(list)
    for r in results:
        by_variant[r.variant].append(r)

    summary = {}
    for variant, vresults in by_variant.items():
        total = len(vresults)

        # Route accuracy
        routable = [r for r in vresults if r.route_correct is not None]
        route_acc = sum(1 for r in routable if r.route_correct) / len(routable) if routable else None

        # Tool validity
        toolable = [r for r in vresults if r.tool_correct is not None]
        tool_acc = sum(1 for r in toolable if r.tool_correct) / len(toolable) if toolable else None

        # Answer contains
        containable = [r for r in vresults if r.contains_total > 0]
        if containable:
            total_hits = sum(r.contains_hits for r in containable)
            total_expected = sum(r.contains_total for r in containable)
            answer_contains = total_hits / total_expected if total_expected else None
        else:
            answer_contains = None

        # Answer excludes
        excludable = [r for r in vresults if r.excludes_total > 0]
        if excludable:
            total_violations = sum(r.excludes_hits for r in excludable)
            total_checks = sum(r.excludes_total for r in excludable)
            answer_excludes = 1 - (total_violations / total_checks) if total_checks else None
        else:
            answer_excludes = None

        # Avg latency
        avg_latency = sum(r.latency_ms for r in vresults) / total if total else 0

        summary[variant] = {
            "total_cases": total,
            "route_accuracy": route_acc,
            "tool_validity": tool_acc,
            "answer_contains": answer_contains,
            "answer_excludes": answer_excludes,
            "avg_latency_ms": round(avg_latency),
        }

    return summary


def print_summary_table(summary: dict) -> None:
    """Print a nice console table."""
    print()
    print("=" * 90)
    print("  TABLE 6: Graph & Agent Ablation Results")
    print("=" * 90)
    print(f"  {'Variant':<25} {'Route Acc':>10} {'Tool Valid':>11} {'Ans Contains':>13} {'Ans Excludes':>13} {'Avg ms':>8}")
    print("-" * 90)

    for variant, metrics in summary.items():
        route = f"{metrics['route_accuracy']:.1%}" if metrics['route_accuracy'] is not None else "n/a"
        tool = f"{metrics['tool_validity']:.1%}" if metrics['tool_validity'] is not None else "n/a"
        contains = f"{metrics['answer_contains']:.1%}" if metrics['answer_contains'] is not None else "n/a"
        excludes = f"{metrics['answer_excludes']:.1%}" if metrics['answer_excludes'] is not None else "n/a"
        latency = f"{metrics['avg_latency_ms']}"

        print(f"  {variant:<25} {route:>10} {tool:>11} {contains:>13} {excludes:>13} {latency:>8}")

    print("=" * 90)
    print()


def print_detailed_results(results: list[CaseResult], variant: str) -> None:
    """Print per-case results for a variant."""
    vresults = [r for r in results if r.variant == variant]
    if not vresults:
        return

    print(f"\n{'─' * 70}")
    print(f"  Detailed: {variant} ({len(vresults)} cases)")
    print(f"{'─' * 70}")

    for r in vresults:
        route_icon = "✅" if r.route_correct else ("❌" if r.route_correct is False else "⬜")
        ans_icon = "✅" if (r.contains_pass and r.excludes_pass) else "❌" if r.contains_total > 0 else "⬜"

        print(f"  [{r.case_id:>2}] {route_icon} route={r.actual_route:<12} "
              f"{ans_icon} ans={r.contains_hits}/{r.contains_total} "
              f"excl={r.excludes_hits}/{r.excludes_total} "
              f"[{r.latency_ms}ms] {r.category}")
        if r.error:
            print(f"       ⚠️  {r.error[:80]}")


def export_json(results: list[CaseResult], summary: dict, output_path: Path) -> None:
    """Export results to JSON."""
    output_path.parent.mkdir(parents=True, exist_ok=True)
    data = {
        "summary": summary,
        "results": [asdict(r) for r in results],
    }
    output_path.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
    logger.info("📄 Results exported to %s", output_path)


def export_excel(results: list[CaseResult], summary: dict, output_path: Path) -> None:
    """Export results to Excel."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        logger.warning("openpyxl not installed. Skipping Excel export.")
        return

    wb = openpyxl.Workbook()

    # === Summary Sheet ===
    ws_summary = wb.active
    ws_summary.title = "Summary"

    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    pass_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    fail_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    na_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

    headers = ["Variant", "Cases", "Route Accuracy", "Tool Validity",
               "Answer Contains", "Answer Excludes", "Avg Latency (ms)"]
    for c, h in enumerate(headers, 1):
        cell = ws_summary.cell(row=1, column=c, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for r, (variant, metrics) in enumerate(summary.items(), 2):
        ws_summary.cell(row=r, column=1, value=variant)
        ws_summary.cell(row=r, column=2, value=metrics["total_cases"])

        for c, key in enumerate(["route_accuracy", "tool_validity", "answer_contains", "answer_excludes"], 3):
            val = metrics[key]
            cell = ws_summary.cell(row=r, column=c)
            if val is None:
                cell.value = "n/a"
                cell.fill = na_fill
            else:
                cell.value = val
                cell.number_format = "0.0%"
                cell.fill = pass_fill if val >= 0.8 else fail_fill

        ws_summary.cell(row=r, column=7, value=metrics["avg_latency_ms"])

    for col in ws_summary.columns:
        ws_summary.column_dimensions[col[0].column_letter].width = 18

    # === Details Sheet ===
    ws_detail = wb.create_sheet("Details")
    detail_headers = ["Case ID", "Variant", "Category", "Query",
                      "Expected Route", "Actual Route", "Route OK",
                      "Contains Hits", "Contains Total", "Contains OK",
                      "Excludes Violations", "Excludes Total", "Excludes OK",
                      "Latency (ms)", "Response (excerpt)", "Error"]

    for c, h in enumerate(detail_headers, 1):
        cell = ws_detail.cell(row=1, column=c, value=h)
        cell.font = header_font
        cell.fill = header_fill

    for r_idx, result in enumerate(results, 2):
        ws_detail.cell(row=r_idx, column=1, value=result.case_id)
        ws_detail.cell(row=r_idx, column=2, value=result.variant)
        ws_detail.cell(row=r_idx, column=3, value=result.category)
        ws_detail.cell(row=r_idx, column=4, value=result.query[:80])
        ws_detail.cell(row=r_idx, column=5, value=result.expected_route)
        ws_detail.cell(row=r_idx, column=6, value=result.actual_route)
        route_cell = ws_detail.cell(row=r_idx, column=7)
        if result.route_correct is None:
            route_cell.value = "n/a"
        else:
            route_cell.value = "✅" if result.route_correct else "❌"
        ws_detail.cell(row=r_idx, column=8, value=result.contains_hits)
        ws_detail.cell(row=r_idx, column=9, value=result.contains_total)
        ws_detail.cell(row=r_idx, column=10, value="✅" if result.contains_pass else ("❌" if result.contains_total > 0 else "n/a"))
        ws_detail.cell(row=r_idx, column=11, value=result.excludes_hits)
        ws_detail.cell(row=r_idx, column=12, value=result.excludes_total)
        ws_detail.cell(row=r_idx, column=13, value="✅" if result.excludes_pass else "❌")
        ws_detail.cell(row=r_idx, column=14, value=result.latency_ms)
        ws_detail.cell(row=r_idx, column=15, value=result.response[:200])
        ws_detail.cell(row=r_idx, column=16, value=result.error[:100])

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_path))
    logger.info("📊 Excel exported to %s", output_path)


# ─────────────────────────────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────────────────────────────

VARIANT_RUNNERS = {
    "fixed_route": lambda cases: run_fixed_route(cases),
    "supervisor": lambda cases: asyncio.get_event_loop().run_until_complete(run_supervisor(cases)),
    "full_system": lambda cases: asyncio.get_event_loop().run_until_complete(run_full_system(cases)),
}


def main():
    parser = argparse.ArgumentParser(description="Ablation Test — Table 6")
    parser.add_argument("--variant", choices=list(VARIANT_RUNNERS.keys()),
                        help="Run a specific variant")
    parser.add_argument("--all", action="store_true", help="Run all variants")
    parser.add_argument("--dry-run", action="store_true", help="Only run 3 cases")
    parser.add_argument("--export-excel", action="store_true", help="Export results to Excel")
    parser.add_argument("--export-json", action="store_true", help="Export results to JSON")
    parser.add_argument("--verbose", action="store_true", help="Show detailed per-case results")
    args = parser.parse_args()

    if not args.variant and not args.all:
        parser.error("Specify --variant <name> or --all")

    limit = 3 if args.dry_run else None
    cases = load_dataset(limit=limit)

    variants = list(VARIANT_RUNNERS.keys()) if args.all else [args.variant]
    all_results: list[CaseResult] = []

    for variant in variants:
        logger.info("=" * 60)
        logger.info("🏃 Running variant: %s (%d cases)", variant, len(cases))
        logger.info("=" * 60)

        t0 = time.time()
        runner = VARIANT_RUNNERS[variant]
        results = runner(cases)
        elapsed = time.time() - t0

        all_results.extend(results)
        logger.info("✅ %s completed in %.1fs", variant, elapsed)

        if args.verbose:
            print_detailed_results(all_results, variant)

    # Summary
    summary = summarize(all_results)
    print_summary_table(summary)

    # Exports
    if args.export_json or args.all:
        json_path = RESULTS_DIR / "ablation_results.json"
        export_json(all_results, summary, json_path)

    if args.export_excel:
        xlsx_path = RESULTS_DIR / "ablation_results.xlsx"
        export_excel(all_results, summary, xlsx_path)

    # Print total
    total_pass = sum(1 for r in all_results if r.route_correct is True)
    total_route = sum(1 for r in all_results if r.route_correct is not None)
    print(f"Overall route accuracy: {total_pass}/{total_route} "
          f"({total_pass / total_route:.1%})" if total_route else "No route tests.")


if __name__ == "__main__":
    main()
